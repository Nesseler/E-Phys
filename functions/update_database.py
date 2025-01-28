# -*- coding: utf-8 -*-
"""
Created on Tue Jan 28 14:41:04 2025

@author: nesseler
"""

import pandas as pd
from openpyxl import workbook 
from openpyxl import load_workbook

def update_analyzed_sheet(analyzed_cell_IDs, PGF = 'cc_rest'):
    '''
    This function loads and updates the analyzed sheet in the ePhys-database.
    '''

    from parameters.directories_win import table_file
    
    # load analyzed sheet with pandas for easy indexing
    analyzed = pd.read_excel(table_file,
                             sheet_name = 'analyzed',
                             index_col = 'cell_ID')
    
    # load database excel file
    ePhys_workbook = load_workbook(table_file)
    
    # define analysed sheet
    analyzed_sheet = ePhys_workbook['analyzed']
    
    # iterate through cell_IDs
    for cell_ID in analyzed_cell_IDs:
        # get row and column index from analyzed worksheet loaded with pandas
        row = analyzed.index.get_loc(cell_ID)
        
        # check for multiple rows per cell
        if type(row) == slice:
            row = row.start
        
        # add for indices
        row = row + 2   # for 1-based excel indexing and column header
        
        col = analyzed.columns.get_loc(PGF) + 2     # for 1-based excel indexing and index
        
        # write to cell
        analyzed_sheet.cell(row = row, column = col).value = 1
    
    # save workbook
    ePhys_workbook.save(table_file)
    
    # close file
    ePhys_workbook.close()

    del ePhys_workbook
    del analyzed_sheet
    
    import gc
    gc.collect()
    
    # from time import sleep
    # sleep(0.5)
    # yield ePhys_workbook
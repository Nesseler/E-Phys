# -*- coding: utf-8 -*-
"""
Created on Fri May  2 15:07:06 2025

@author: nesseler
"""

import matplotlib.pyplot as plt
import matplotlib as mtl
from matplotlib.patches import Rectangle
import numpy as np
import pandas as pd
import scipy as sc
# import sys

from miniML import MiniTrace, EventDetection
from miniML_plot_functions import miniML_plots


# %% generall functions

def calc_PDF_fromTimepoints(data: np.array, bins: np.array):

    # calc inter-event interval (IEI)
    IEI = np.diff(data, prepend = 0)    

    # calc mean rate
    mean_rate = 1 / np.mean(IEI)
    
    # calculate the probability density function with bins given
    pdf = mean_rate * np.exp(- mean_rate * bins)

    return pdf


def calc_CDF_fromTimepoints(data: np.array, bins: np.array):
    
    pdf = calc_PDF_fromTimepoints(data, bins)

    cdf = [np.max(pdf) - p for p in pdf]

    return cdf


def minmax_norm(data: np.array, minzero: bool = False):
    
    # calc min max normalized version of array 
    if minzero:
        normed = (data - 0) / (np.max(data) - 0)
    else:
        normed = (data - np.min(data)) / (np.max(data)- np.min(data))
    
    return normed


def round_up_to_base(number, base):
    return base * np.ceil(number/base)  


def round_down_to_base(number, base):
    return base * np.floor(number/base)


def create_uniform_events(n_events: int = 100, t: float = 180) -> np.array:
    # uniformly distribute the same number of events as measured
    # by randomly draw from uniform distribution between t0 and t_end
    uni_events = sc.stats.uniform.rvs(loc = 0, scale = t, size = n_events)
    
    # sort values to perserve time series aspect
    uni_events = np.sort(uni_events)
        
    return uni_events


# %% plotting functions

# init plotting
mtl.rcParams.update({'font.size': 9, 'font.family' : 'Arial'})


def trace_events_prediction(t, trange):

    # filter prediction
    filtered_prediction = sc.ndimage.maximum_filter1d(detection.prediction, size=int(5*detection.interpol_factor), origin=-2)

    # fill with nan values
    ext_prediction = np.append(filtered_prediction, [np.nan]*(data_filtered.shape[0]-filtered_prediction.shape[0]))
    
    # init figure
    parent_fig = plt.figure(layout='constrained',
                            figsize = (6, 3),
                            dpi=300)
    
    # set figure title
    parent_fig.suptitle(f'{cell_ID} - {recording} - Prediction - Trace - Events',
                        fontsize = 9)
    
    
    x_full = np.arange(0, data_filtered.shape[0] / SR, step=1/SR)
    
    plt_idc = np.arange((0+t) * SR, (trange+t) * SR, step = 1, dtype = int)
    plt_peak_idc = [idx for idx in detection.event_peak_locations if (idx > plt_idc[0] and idx < plt_idc[-1])]
    
    axs = parent_fig.subplots(nrows=3,
                              ncols=1,
                              sharex=True,
                              height_ratios=[1, 0.3, 2])
    
    # plot prediction
    axs[0].plot(x_full[plt_idc], ext_prediction[plt_idc],
                color='k',
                alpha=0.5,
                lw=0.5,
                label='filtered')
    
    axs[0].hlines(xmin = x_full[plt_idc][0], xmax = x_full[plt_idc][-1],
                  y = model_th, 
                  color = 'r',
                  lw = 0.5,
                  ls = 'dashed')
    
    # plot eventplot
    axs[1].eventplot(positions = x_full[plt_peak_idc],
                      orientation = 'horizontal',
                      lineoffsets = 0,
                      linelengths = 1,
                      linewidths = 1,
                      color = 'r',
                      label = 'events')
    
    # plot data
    axs[2].plot(x_full[plt_idc], data_filtered[plt_idc],
                color='k',
                lw=0.5,
                label='data (filt)',
                zorder = 2)
    
    # plot event peak indicators
    axs[2].scatter(x = x_full[plt_peak_idc],
                    y = data_filtered[plt_peak_idc],
                    marker = 'o',
                    color = 'r',
                    s = 5,
                    lw = 1,
                    zorder = 1)
    
    
    # remove spines
    [ax.spines[spine].set_visible(False) for ax in axs for spine in ['top', 'right']]
    
    # axis 
    axs[0].set_ylim([-0.05, 1.05])
    axs[0].set_ylabel('Probability')
    
    axs[1].set_ylim([-1, 1])
    axs[1].set_ylabel('Events')
    axs[1].set_yticks(ticks = [0], labels = [])
    
    axs[2].set_xlim([0+t-(trange*0.01), trange+t+(trange*0.01)])
    [ax.spines['bottom'].set_bounds([0+t, trange+t]) for ax in axs]
    axs[2].set_xlabel('Time [s]')
    
    axs[2].hlines(xmin = 0, xmax = data_filtered.shape[0], y = 0, zorder = 0, lw = 0.5, color = 'k', ls = 'dashed', alpha = 0.5)
    
    trace_mean = np.mean(data_filtered)
    ax_min = int(round_down_to_base(trace_mean-40, 10))
    ax_max = int(round_up_to_base(trace_mean+10, 10))
    ax_range = ax_max - ax_min
    stepminor = 5

    axs[2].set_ylim([ax_min-(ax_range*0.01), ax_max+(ax_range*0.01)])
    axs[2].spines['left'].set_bounds([ax_min, ax_max])
    axs[2].set_yticks(ticks = [ax_min, 0, ax_max], labels = [ax_min, '', ax_max])
    axs[2].set_yticks(ticks = np.arange(ax_min, ax_max+(stepminor/10), stepminor), minor = True)
    axs[2].set_ylabel('Current [pA]')
    
    parent_fig.align_labels()
    
    plt.show()

    # save figure
    parent_fig.savefig(vplots_path + '/vc_PSCs-trace_events_prediction/' + f'{cell_ID}-{recording}-trace_events_prediction' + ".png", 
                        format = 'png')


def amplitude_histogram():
    # set data and bins
    data = detection.event_stats.amplitudes
    bins = np.arange(-50, 0+1, 1)
    
    # init figure
    fig, ax = plt.subplots(1,1, dpi = 300, layout = 'constrained')
    
    # set title
    fig.suptitle(f'{cell_ID} {recording} amplitude histogram')
    ax.hist(data, bins=bins, 
            histtype='step', 
            color='k',
            lw = 1)
    
    # edit axis
    ax.set_ylabel('Count [#]')
    ax.set_ylim([-2, 132])
    ax.set_yticks(ticks = np.arange(0, 130+1, 20))
    ax.set_yticks(ticks = np.arange(0, 130+1, 2), minor = True)
    ax.spines['left'].set_bounds([0, 130])
    
    ax.set_xlabel('Amplitude [pA]')
    ax.set_xticks(ticks = np.arange(-50, 0+1, 10))
    ax.set_xticks(ticks = np.arange(-50, 0, 1), minor = True)
    ax.set_xlim([-51, 1])
    ax.spines['bottom'].set_bounds([-50, 0])
    
    # remove spines
    [ax.spines[spine].set_visible(False) for spine in ['top', 'right']]
    
    # display figure
    plt.show()
    
    # save figure
    fig.savefig(vplots_path + '/vc_PSCs-ampl_hist/' + f'{cell_ID}-{recording}-ampl_hist' + ".png", 
                  format = 'png')


def average_event():
    # set data
    events = detection.events[detection.singular_event_indices]
    event_x = np.arange(0, events.shape[1]) * 1/(SR/1e3)
    event_average = np.mean(events, axis=0)
    
    # init figure
    fig, ax = plt.subplots(1, 1, dpi = 300, layout = 'constrained')
    
    # set title
    fig.suptitle(f'{cell_ID} {recording} average event')
        
    # plot
    ax.plot(event_x, events.T,
            c = 'k',
            alpha = 0.3,
            lw = 1,
            label = '_nolegend_')
    
    ax.plot(event_x, event_average,
            c = 'r', 
            lw = 1.5,
            label = 'average event')
    
    plt.legend(frameon = False,
               fontsize = 7,
               loc = 'lower right')
    
    # edit axis
    ax.set_ylabel('Current [pA]')
    ax.set_ylim([-51, 10])
    ax.set_yticks(ticks = np.arange(-50, 10+1, 10))
    ax.set_yticks(ticks = np.arange(-50, 10+1, 2), minor = True)
    ax.spines['left'].set_bounds([-50, 10])
    
    ax.set_xlabel('Time [ms]')
    ax.set_xticks(ticks = np.arange(0, round(event_x[-1])+1, 20))
    ax.set_xticks(ticks = np.arange(0, round(event_x[-1])+1, 5), minor = True)
    ax.set_xlim([-5, round(event_x[-1])+5])
    ax.spines['bottom'].set_bounds([0, round(event_x[-1])])
    
    # remove spines
    [ax.spines[spine].set_visible(False) for spine in ['top', 'right']]
    
    # display figure
    plt.show()
    
    # save figure
    fig.savefig(vplots_path + '/vc_PSCs-average_event/' + f'{cell_ID}-{recording}-average_event' + ".png", 
                format = 'png')
    
    
def event_frequency():
    # for figure
    trace = data_filtered
    event_locations = detection.event_peak_locations
    x = np.arange(0, data_filtered.shape[0] / SR, 1/SR)
    bins = np.arange(0, 30+0.05, 0.05)
    
    n_events = event_locations.shape[0]
    uniform_events = create_uniform_events(n_events, t=trace.shape[0] / SR)
    
    x_max = trace.shape[0] / SR
    trace_mean = np.mean(trace)
    
    events_hist, _ = np.histogram(np.diff(event_locations / SR, prepend = 0), bins = bins)
    uniform_hist, _ = np.histogram(np.diff(uniform_events, prepend = 0), bins = bins)
    theo_pdf = calc_PDF_fromTimepoints(data = uniform_events, bins = bins)
    theo_cdf = calc_CDF_fromTimepoints(data = uniform_events, bins = bins)
    
    
    # init figure
    fig, axs = plt.subplots(nrows = 3,
                            ncols = 1, 
                            dpi = 300, layout = 'constrained', 
                            height_ratios = [2, 1, 7])
    
    # figure title
    fig.suptitle(f'{cell_ID} {recording} - Event frequency (IEI)')
    
    # trace
    ax = axs[0]
    
    # title
    # ax.set_title(axs_titles[ti], loc = 'left', fontsize = 9)
    
    # plot prediction
    ax.plot(x, trace,
            color = 'k',
            lw=0.5,
            label= 'trace', 
            zorder = 2)
    
    ax.scatter(x[event_locations], trace[event_locations],
               color = 'r',
               alpha = 0.5,
               s = 3,
               lw=0.5,
               label='events',
               zorder = 1)
    
    ax.legend(frameon = False,
              fontsize = 9,
              loc = 'lower right',
              labelspacing = 0.1,
              borderpad = 0.1,
              borderaxespad = 0.1)
    
    ax.hlines(xmin = 0, xmax = x_max, y = 0, zorder = 0, lw = 0.5, color = 'k', ls = 'dashed', alpha = 0.5)
    
    ax_min = int(round_down_to_base(trace_mean-40, 10))
    ax_max = int(round_up_to_base(trace_mean+10, 10))
    ax_range = ax_max - ax_min
    stepminor = 5

    ax.set_ylim([ax_min-(ax_range*0.01), ax_max+(ax_range*0.01)])
    ax.spines['left'].set_bounds([ax_min, ax_max])
    ax.set_yticks(ticks = [ax_min, 0, ax_max], labels = [ax_min, '', ax_max])
    ax.set_yticks(ticks = np.arange(ax_min, ax_max+(stepminor/10), stepminor), minor = True)
    ax.set_ylabel('Current [pA]')
    

    # eventplots
    ax = axs[1]
    
    ax.eventplot(positions = x[event_locations],
                  orientation = 'horizontal',
                  lineoffsets = 2,
                  linelengths = 1,
                  linewidths = 0.5,
                  color = 'k',
                  label = 'events')
    
    ax.eventplot(positions = uniform_events,
                  orientation = 'horizontal',
                  lineoffsets = 0,
                  linelengths = 1,
                  linewidths = 0.5,
                  color = 'gray',
                  alpha = 0.5,
                  label = 'uniform')
    
    for ax in axs[:2]:
        ax.set_xlim([0-1.8, x_max+1.8])
        ax.spines['bottom'].set_bounds([0, x_max])
        ax.set_xticks(ticks = np.arange(0, x_max+1, 60),
                            labels = [])
        ax.set_xticks(ticks = np.arange(0, x_max+1, 10), minor = True)
    
    
    # remove_spines_n_ticks([axs[0]], axis = 'x')
    axs[1].set_xticks(ticks = np.arange(0, x_max+1, 60),
                      labels = np.arange(0, x_max+1, 60, dtype = int))
    axs[1].set_xlabel('Time [s]')
    
    axs[1].set_yticks(ticks = [0, 2], labels = ['uniform', 'events'])
    axs[1].spines['left'].set_visible(False)
    
    
    # histogram
    ax = axs[2]
    
    ax.stairs(minmax_norm(events_hist), bins, 
              fill = False,
              lw = 1,
              color = 'k',
              label = 'IEI')
    
    ax.plot(bins, minmax_norm(theo_pdf),
            color = 'k',
            lw = 1,
            ls = 'dashed',
            label = 'uniform pdf')
    
    ax.plot(bins, minmax_norm(theo_cdf),
            color = 'k',
            lw = 1,
            ls = 'dotted',
            label = 'uniform cdf')
    
    ax.legend(frameon = False, loc = 'lower right', fontsize = 9)
    
    ax.set_xlabel('IEI [s]')
    ax.set_xticks(ticks = np.arange(0, 5+0.1, 1))
    ax.set_xticks(ticks = np.arange(0, 5+0.1, 0.2), minor = True)
    ax.set_xlim([0-0.1, 5+0.1])
    ax.spines['bottom'].set_bounds([0, 5])
    
    ax.set_ylabel('Normalized event count')
    ax.set_yticks(ticks = np.arange(0, 1+0.01, 0.2))
    ax.set_yticks(ticks = np.arange(0, 1+0.01, 0.05), minor = True)
    ax.set_ylim([0-0.01, 1+0.01])
    ax.spines['left'].set_bounds([0, 1])
    
    
    # inset axis: full histogram
    ax_inset = ax.inset_axes([3/5, 0.55, 1.8/5, 0.35])
    
    ax_inset.stairs(events_hist, bins, 
                    fill = False,
                    lw = 1,
                    color = 'k',
                    label = 'IEI')
    
    # inset marker
    box_xmin   = 0
    box_width  = 5 
    box_ymin   = 0
    box_height = round_up_to_base(np.max(events_hist), 5)
    
    # add rectangle marker
    ax_inset.add_patch(Rectangle(xy = (box_xmin, box_ymin), 
                       width = box_width, 
                       height = box_height,
                       fill = False,
                       color = 'r',
                       linestyle = '--',
                       lw = 0.5))
    
    ax_inset.set_ylabel('Count [#]')
    
    ax_inset.set_xlabel('IEI [s]')
    ax_inset.set_xticks(ticks = np.arange(0, bins[-1]+0.1, 10))
    ax_inset.set_xticks(ticks = np.arange(0, bins[-1]+0.1, 1), minor = True)
    ax_inset.set_xlim([0-0.1, bins[-1]+0.1])
    ax_inset.spines['bottom'].set_bounds([0, bins[-1]])
    [ax_inset.spines[spine].set_visible(False) for spine in ['top', 'right']]
    
    # remove spines
    [ax.spines[spine].set_visible(False) for ax in axs for spine in ['top', 'right']]
    
    # display figure
    plt.show()
    
    # save figure
    fig.savefig(vplots_path + '/vc_PSCs-inter_event_intervals/' + f'{cell_ID}-{recording}-IEI' + ".png", 
                format = 'png')
    


# %% export functions    
    
from openpyxl import workbook 
from openpyxl import load_workbook

def update_analyzed_sheet(analyzed_cell_IDs, PGF = 'cc_rest'):
    '''
    This function loads and updates the analyzed sheet in the ePhys-database.
    '''
    
    # load analyzed sheet with pandas for easy indexing
    analyzed = pd.read_excel(ePhys_parent + '/ePhys-database.xlsx',
                             sheet_name = 'analyzed',
                             index_col = 'cell_ID')
    
    # load database excel file
    ePhys_workbook = load_workbook(ePhys_parent + '/ePhys-database.xlsx')
    
    # define analysed sheet
    analyzed_sheet = ePhys_workbook['analyzed']
    
    # iterate through cell_IDs
    for cell_ID in analyzed_cell_IDs:
        # get row and column index from analyzed worksheet loaded with pandas
        row = analyzed.index.get_loc(cell_ID)
        
        # check for multiple rows per cell
        if type(row) == slice:
            row = row.start
        
        # add for indices
        row = row + 2   # for 1-based excel indexing and column header
        
        col = analyzed.columns.get_loc(PGF) + 2     # for 1-based excel indexing and index
        
        # write to cell
        analyzed_sheet.cell(row = row, column = col).value = 1
    
    # save workbook
    ePhys_workbook.save(ePhys_parent + '/ePhys-database.xlsx')
    
    # close file
    ePhys_workbook.close()

    del ePhys_workbook
    del analyzed_sheet
    
    import gc
    gc.collect()
    

# %% setup paths

# paths
ePhys_parent = 'Z:/n2021_MOS_AOS_Integration/ePhys-BAOT_MeA'
rawData_path = 'Z:/n2021_MOS_AOS_Integration/ePhys-BAOT_MeA/RAW_data/'
data_path = 'Z:/n2021_MOS_AOS_Integration/ePhys-BAOT_MeA/synaptic_currents/'
vplots_path = 'Z:/n2021_MOS_AOS_Integration/ePhys-BAOT_MeA/vplots'

# %% setup parameters

# load
eholds = ['Erest', 'Ek'] #'Ezero', 'Ep30', 'ECl'
treatments = ['ctrl', 'AP5_NBQX', 'AP5_NBQX_washin', 'GBZ', 'GBZ_washin', 'AP5_NBQX_GBZ', 'AP5_NBQX_GBZ_washin', 'adaEk']
times = ['3min']

# PGF 
scaling = 1e12
unit = 'pA'
SR = 100_000
total_dur = 180
sweep_dur = 30

# analysis
filter_freq = 750
filter_order = 3
factor = 19
model_th = 0.5
winsize_ms = int((600 * factor) / (SR/1e3))

# output
plot = True
save = True

# set settings for event detection
eventdtc = {'direction' : 'negative',
            'window_size' : 600 * factor,
            'model_threshold' : 0.5,
            'batch_size' : 512,
            'event_detection_peakw' : 5,
            'stride' : 30,
            'rel_prom_cutoff' : 0.25,
            'convolve_win' : 20 * factor}


# %% load lookups

# load xlsx sheet
lookup = pd.read_excel(ePhys_parent + '/ePhys-database.xlsx', 
                       sheet_name = 'PGFs_Syn', 
                       index_col = 'cell_ID')

# load analyzed sheet
analyzed = pd.read_excel(ePhys_parent + '/ePhys-database.xlsx',
                         sheet_name = 'analyzed',
                         index_col = 'cell_ID')

# %% load data

time = times[0]

# setup string
newly_analyzed_cellIDs = str()

for treat in treatments:
    for ehold in eholds:
                
        # set recording type as named in PatchMaster
        rectype = f'vc-{ehold}-{time}'
        
        # set recording as named in xlsx sheet
        recording = f'vc-{ehold}-{time}-{treat}'
        
        # get cell_IDs
        try:
            cell_IDs = lookup.loc[:, recording].dropna().index.to_list()
        
        except KeyError:
            cell_IDs = []
        
        # init analyzed cell_IDs
        analyzed_cell_IDs = list()
        
        # iterate through cell_IDs
        for cell_ID in cell_IDs:
            
            # check if already analyzed
            if not np.isnan(analyzed.at[cell_ID, recording]).any():
                analyzed_cell_IDs.append(cell_ID)
            
            else:   
                # get cell row
                cell_lookup = lookup.loc[cell_ID, ['file', 'group', recording, (recording + '-steps')]]
                
                # special case for cells in two files
                if len(cell_lookup.shape) > 1:
                    cell_lookup = cell_lookup.dropna().iloc[0]
                
                # get group + series index & filename
                group_idx = cell_lookup['group']
                series_idx = cell_lookup[recording] -1
                rawData_filename = cell_lookup['file'] + '.dat'
                
                # load trace to MiniTrace
                trace = MiniTrace.from_heka_file(filename = (rawData_path + rawData_filename),
                                                 rectype = rectype,
                                                 group = group_idx,
                                                 exclude_series = np.delete(np.arange(0, 100), int(series_idx)),
                                                 scaling = scaling,
                                                 unit = unit)
                
                # filter trace
                b, a = sc.signal.bessel(filter_order, filter_freq, fs = SR)
                data_filtered = sc.signal.lfilter(b, a, trace.data)
            
                # get sweep to be included
                sweep_idc = [int(s) for s in cell_lookup[recording + '-steps'].split(',')]
                
                # set first and last index
                first_idx = 0
                last_idx = int(total_dur*SR)
                
                # keep all
                if sweep_idc == [1,2,3,4,5,6]:
                    pass
                
                # remove first sweeps
                elif (sweep_idc[0] != 1) and not (sweep_idc[-1] != 6):
                    first_idx = int((sweep_idc[0]-1)*sweep_dur*SR)
            
                # remove last sweeps
                elif not (sweep_idc[0] != 1) and (sweep_idc[-1] != 6):
                    last_idx = int(sweep_idc[-1]*sweep_dur*SR)
                    
                # remove first and last sweeps
                elif (sweep_idc[0] != 1) and (sweep_idc[-1] != 6):
                    first_idx = int((sweep_idc[0]-1)*sweep_dur*SR)
                    last_idx = int(sweep_idc[-1]*sweep_dur*SR)
                    
                elif np.max(sweep_idc) > 6:
                    raise ValueError
                
                # apply
                data_filtered = data_filtered[first_idx:last_idx]
                
                # create new miniTrace object
                trace = MiniTrace(data = data_filtered,
                                  sampling_interval = 1 / SR,
                                  y_unit = unit,
                                  filename = 'None')
                
                
                # %% run event detection
                
                # run prediction
                detection = EventDetection(data=trace,
                                            model_path='C:/Users/nesseler/miniML/models/GC_lstm_model.h5',
                                            window_size = eventdtc['window_size'],
                                            model_threshold = eventdtc['model_threshold'],
                                            batch_size = eventdtc['batch_size'],
                                            event_direction = eventdtc['direction'],
                                            compile_model = True,
                                            verbose = 2)
            
                # detect events
                detection.detect_events(eval = True,
                                        stride = eventdtc['stride'],
                                        peak_w = eventdtc['event_detection_peakw'],
                                        rel_prom_cutoff = eventdtc['rel_prom_cutoff'],
                                        convolve_win = eventdtc['convolve_win'],
                                        resample_to_600 = True)
                
                
                # %% plots
                
                if plot:
                    # full trace + events
                    trace_events_prediction(t = 0, trange = data_filtered.shape[0] / SR)
                             
                    # histogram
                    amplitude_histogram()
            
                    # average event
                    average_event()
                    
                    # event frequency
                    event_frequency()
                                                   
                    
                # %% results export
                
                if save:
                    # convert winsize to string
                    winsize_str = str(int(winsize_ms))
                    
                    # get model threshold
                    th_str = str(model_th).replace('.', 'p')
                    
                    # create filename
                    filename = f'miniMLdetect_{cell_ID}_{ehold}_{time}_{treat}' 
                    
                    # save to pickle file
                    detection.save_to_pickle(filename = data_path + f'miniML_dtc-{ehold}-{treat}/' + filename + '.pickle', 
                                              include_prediction = True, 
                                              include_data = True)
                
                # add to list of analyzed cell_IDs
                analyzed_cell_IDs.append(cell_ID)
                
                # newly analyzed cell_IDs + condition
                newly_analyzed_cellIDs = newly_analyzed_cellIDs + '\n' + f' {cell_ID} {ehold} {treat} {time}'

                
        # %% write to analyzed xlsx sheet
        
        update_analyzed_sheet(analyzed_cell_IDs, PGF = recording)


# %% write txt file

from datetime import datetime

if save:

    # get date and time
    datetime_str = datetime.now().strftime('%Y%m%d_%H%M')
    
    # set path and filename
    settingsoutput_path = data_path + 'miniML_dtc-analysis_settings/'
    txt_filename = 'miniML_dtc-' + datetime_str
    
    with open(settingsoutput_path + txt_filename + '.txt', 'w+') as txt:
        
        # filter
        txt.write(f'Filter\n Filter type: bessel\n Filter freq: {filter_freq}\n Filter order: {filter_order}')
    
        # PGF
        txt.write(f'\n\nPGF\n Scaling: {scaling}\n Unit: {unit}\n SR [Hz]: {SR}\n Total duration [s]: {total_dur}\n Sweep duration [s]: {sweep_dur}')
    
        # analysis
        txt.write(f'\n\nAnalysis:\n Model threshold: {model_th}\n Window size [ms]: {winsize_ms}\n Factor (of original winsize): {factor}')
        
        # analysis dict
        analy_s = '\n\n'
        for k, v in eventdtc.items():
            analy_s = analy_s + ' ' + k.capitalize() + ': ' + str(v) + '\n'
        txt.write(analy_s)
            
        # analyzed cells
        txt.write('\nAnalyzed:' + newly_analyzed_cellIDs)
    
        
# %% end

print('Done!')       
        